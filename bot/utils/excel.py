import openpyxl
import numpy as np
import os

from .data import position_transform, MONTHS
from .module import XLSXCurrentPath

class Excel:
    """A excelfile's controlor."""
    def __init__(self, filename: str=None) -> None:
        self._filename = filename
        self.calendar_dates = np.arange(0, 37, 1)
        
        self.wb = openpyxl.load_workbook(XLSXCurrentPath("templates/base"))
        
    def create_new_file(self, filename: str) -> None:
        """
        Create a new file. 

        Parameters
        ----------
        filename: :type:`str`
            The filename witch the new file named.
        """
        
        self.wb.save(XLSXCurrentPath(filename, create=True))
    
    def save_information(self, month: int, date: int, t1: int, t2: int, nick: str, p: float, filename: str=None) -> None:
        """
        Save a event to the excel file. 

        Parameters
        ----------
        month: :type:`int`
            No description.
        date: :type:`int`
            No description.
        time: :type:`int`
            No description.
        nick: :type:`str`
            No description.
        p: :type:`float`
            No description.
        filename: :type:`str`
            The filename witch will be the target file.
        """
        
        time = f'{t1}:00-{t2}:00'
        if filename is not None:
            self._filename = filename
            
        ws = self.wb[MONTHS[month]]
        c = ws[position_transform(month=month, date=date)]
        if c.value is None:
            ws[position_transform(month=month, date=date)] = f"{time}\n『{nick}』/{p}\n\n"
        else:
            ws[position_transform(month=month, date=date)] = f"{c.value}{time}\n『{nick}』/{p}\n\n"
        self.wb.save(XLSXCurrentPath(self._filename))
        
    def remove_information(self, month: int, date: int, nick: str, filename: str=None) -> bool:
        """
        Remove a event from the excel file. 

        Parameters
        ----------
        month: :type:`int`
            No description.
        date: :type:`int`
            No description.
        nick: :type:`str`
            No description.
        filename: :type:`str`
            The filename witch will be the target file.
            
        Returns
        -------
        status: :type:`bool`
            Is removing successful.
        """
        
        if filename is not None:
            self._filename = filename
            
        wb = openpyxl.load_workbook(XLSXCurrentPath(self._filename))
        ws = wb[MONTHS[month]]
        c = ws[position_transform(month=month, date=date)]
        
        if c.value is None:
            return False
        
        
        if f"『{nick}』" not in c.value:
            return False
            
        events: list = c.value.split("\n")
        e = []
        
        for i in range(len(events)):
            
            if "『" in events[i] and "』" in events[i] and nick not in events[i]:
                
                e.append(events[i-1])
                e.append(events[i])
                
                if i+1 == len(events):
                    try: e.append(events[i+1])
                    except: pass
                    
                else: e.append(events[i+1])
        
        if len(e) == 0:
            new_events = None
        else:
            new_events = "\n".join(e)
                    
        ws[position_transform(month=month, date=date)] = new_events
            
        wb.save(XLSXCurrentPath(self._filename))
        return True
        
    def get_excel_file_information(self) -> tuple[str, str]:
        """
        Get the information of the target file. 
        
        Returns
        -------
        path: :type:`str`
            The path of the target excel file.
        filename: :type:`str`
            The target file name.
        """
        return XLSXCurrentPath(self._filename), self._filename
        
    def change_target_file(self, filename: str) -> tuple[bool, str | None]:
        """
        Change the target file. 

        Parameters
        ----------
        filename: :type:`str`
            The filename witch will be the target file.

        Returns
        -------
        status: :type:`bool`
            Is loading successful.
        old_filename: :type:`str`| :type:`None`
            The filename before this change.
        """
        
        old_filename = self._filename
        
        if os.path.isfile(XLSXCurrentPath(filename)):
            self._filename = filename
            return True, old_filename
        else:
            return False, old_filename
        
    def get_information(self, month: int, date: int, nick: str, filename: str=None) -> list[dict]:
        """
        Get a event information from the excel file. 

        Parameters
        ----------
        month: :type:`int`
            No description.
        date: :type:`int`
            No description.
        nick: :type:`str`
            No description.
        filename: :type:`str`
            The filename witch will be the target file.
            
        Returns
        -------
        information: :type:`list[dict]`
            Included these keys: "nick", "month", "date", "p", "timerange".
        """
        
        if filename is not None:
            self._filename = filename
            
        wb = openpyxl.load_workbook(XLSXCurrentPath(self._filename))
        ws = wb[MONTHS[month]]
        c = ws[position_transform(month=month, date=date)]
        
        events: list[str] = c.value.split("\n")
        
        info: list[dict] = []
        
        for i in range(len(events)):
        
            if "『" in events[i] and "』" in events[i] and nick in events[i]:
                info.append({
                    "nick": nick,
                    "month": month,
                    "date": date,
                    "p": float(events[i].replace(f"『{nick}』/", "")),
                    "timerange": events[i-1]
                })
                
        return info
    
    def get_target_file_name(self):
        return self._filename
        
        
excel = Excel("calendar")